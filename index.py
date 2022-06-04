# -----------------------IMPORTACIONES---------------------------------------------------------------------------------
from fileinput import close
import pandas as pd
from fileinput import close
import tkinter as tk
from ctypes.wintypes import SIZE
from tkinter import *
from tkinter import Tk, PhotoImage, Label, Canvas, ttk, filedialog, messagebox, ttk, Scrollbar, VERTICAL, HORIZONTAL
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
# ----------------------------------------------------------------------------------------------------------------
# Para importar las Listas del archivo FiltroAreas.py
import numpy as np
from FiltroAreas import FiltroOTROS, FiltroPOLITECNICA, FiltroUNIVERSIDAD, searchfor, FiltroTECNOLOGIA
from FiltroAreas import FiltroTECNM

# Para importar las Listas de Areas por carrera del archivo FiltroAreas.py
from FiltroAreas import FiltroAreasComputacional, FiltroAreasSoftware, FiltroAreasInformatica, FiltroAreasInteligencia, FiltroAreasOtros
from FiltroAreas import FiltroAreasSistemasCom, FiltroAreasSistemasInf, FiltroAreasTecnologias, FiltroAreasTecDeLaInformacion, FiltroAreasTelematica
# ----------------------------------------------------------------------------------------------------------------------


# ........VENTANA DE TKINTER......................................................................................................

ventana = Tk()

ventana.title("Pantalla Principal")

ventana.geometry("1000x680")

ventana.resizable(0, 0)
# -------Menu y Sub-menu con la librería Tkinter en Python-----------------------------------------------
def color_amarillo():
    ventana['bg'] = 'yellow'

def color_verde():
    ventana['bg'] = 'green'

def color_azul():
    ventana['bg'] = 'blue'

def color_rojo():
    ventana['bg'] = 'red'
    print('hola')

def color_gris():
    ventana['bg'] = 'gray'
#-----Salir preguntando desde el label    
def mensaje():
    answer=messagebox.askyesno("Salir","¿Desea salir?,confirme..")
    if(answer):
        ventana.destroy()
        
#---...Salir desde la ventana preguntando----------
def on_closing():
    if messagebox.askokcancel("Salir", "¿Desea salir?, Confirme.."):
        ventana.destroy()

ventana.protocol("WM_DELETE_WINDOW", on_closing)

#---------------------------------------------------------
mi_menu = tk.Menu(ventana)
mi_menu.add_command(label='Pestaña 1', command=color_amarillo)
mi_menu.add_command(label='Pestaña 2', command=color_verde)
mi_menu.add_command(label='Pestaña 3', command=color_azul)

mi_dropdown = tk.Menu(ventana)
mi_dropdown.add_command(label='Amarillo', command=color_amarillo)
mi_dropdown.add_command(label='Verde', command=color_verde)
mi_dropdown.add_command(label='Verde', command=color_azul)
mi_dropdown = tk.Menu(mi_menu, tearoff=0)

mi_dropdown.add_command(label='Grafico 1', command=color_rojo)
mi_dropdown.add_command(label='Grafico 2', command=color_gris)

mi_menu.add_cascade(label='Graficos', menu=mi_dropdown)
mi_menu.add_command(label='salir',command=mensaje)
ventana.config(menu=mi_menu)
#


# ------IMAGENES-----------------------------------------------------------------------------------------
bg = PhotoImage(file="Imagenes/data.png")
# Create Canvas
canvas1 = Canvas(ventana, width=700,

                 height=700)

canvas1.pack(fill="both", expand=True)

# Display image

canvas1.create_image(0, 0, image=bg,

                     anchor="nw")

# .....DEF DE LOS PRIMEROS FILTROS..................................................................................


def abrir_archivo():

    archivo = filedialog.askopenfilename(initialdir='/',

                                         title='Selecione archivo',

                                         filetype=(('xlsx files', '*.xlsx*'), ('All files', '*.*')))

    indica['text'] = archivo


# ------------------------------------------------------------------------------------------------------------------
def datos_excel():

    datos_obtenidos = indica['text']

    archivoexcel = r'{}'.format(datos_obtenidos)

    path = archivoexcel

    libro = openpyxl.load_workbook(path, data_only=True)

    Hoja1 = libro.active
    libro.save("filtro01.xlsx")
# -------------INICIO DEL PRIMER FILTRO-------------------------------------------------------------------------
    # Fijar hoja

    print("Insertando datos en la columna ENTIDAD FEDERATIVA :)")
    renglon = 0

    for filas in Hoja1.iter_rows(1, Hoja1.max_row):

        columna = 0

        renglon = renglon + 1

        if renglon >= 5:

            for celdas in filas:

                columna = columna + 1

                if columna == 1:

                    if celdas.value:

                        estado = celdas.value
                        Hoja1.cell(row=renglon, column=1).value = "***"

                    else:

                        Hoja1.cell(row=renglon, column=1).value = estado

                        print(celdas.value)

    libro.save("Filtro01.xlsx")

# ----------------------------------------------------------------------------------------------


def municipio_excel():

    print("Insertando datos en la columna MUNICIPIO :)")
    # path=archivoexcel

    libro = openpyxl.load_workbook("Filtro01.xlsx")
    Hoja1 = libro.active

    print("Municipios")
    renglon = 0
    for filas in Hoja1.iter_rows(1, Hoja1.max_row):
        columna = 0
        renglon = renglon + 1
        if renglon >= 6:
            for celdas in filas:
                columna = columna + 1
                if columna == 2:
                    if celdas.value:
                        estado = celdas.value
                        Hoja1.cell(row=renglon, column=2).value = "***"
                    else:
                        Hoja1.cell(row=renglon, column=2).value = estado
                        print(celdas.value)

    libro.save("Filtro01.xlsx")

# -------------------------------------------------------------------------------------------------------------


def institucion_excel():
    print(" Insertando datos en la columna INSTITUCIÓN DE EDUCACIÓN SUPERIOR :)")
    book = openpyxl.load_workbook("Filtro01.xlsx")
    Hoja1 = book.active

    renglon = 0
    for filas in Hoja1.iter_rows(1, Hoja1.max_row):
        columna = 0
        renglon = renglon + 1

        if renglon >= 7:
            for celdas in filas:
                columna = columna + 1

                if columna == 3:
                    if celdas.value:
                        estado = celdas.value
                        Hoja1.cell(row=renglon, column=3).value = "***"

                    else:
                        Hoja1.cell(row=renglon, column=3).value = estado
                        print(celdas.value)

    book.save("Filtro01.xlsx")
 # ----------------------------------------------------------------------------------------------------------


def programa_excel():
    print(" Insertando datos en la columna PROGRAMA DE ESTUDIOS  :")
    book = openpyxl.load_workbook("filtro01.xlsx")
    Hoja1 = book.active

    renglon = 0
    for filas in Hoja1.iter_rows(1, Hoja1.max_row):
        columna = 0
        renglon = renglon + 1

        if renglon >= 8:
            for celdas in filas:
                columna = columna + 1

                if columna == 4:
                    if celdas.value:
                        municipio = celdas.value

                        Hoja1.cell(row=renglon, column=4).value = municipio
                        print(celdas.value)
                    else:
                        Hoja1.cell(row=renglon, column=4).value = "***"

    Hoja1.delete_rows(idx=1, amount=4)
    # Hoja1.delete_cols(5,10)

    Hoja1.cell(row=1, column=1, value='ENTIDAD FEDERATIVA')
    Hoja1.cell(row=1, column=2, value='MUNICIPIO')
    Hoja1.cell(row=1, column=3, value='INSTITUCIÓN DE EDUCACIÓN SUPERIOR')
    Hoja1.cell(row=1, column=4, value='PROGRAMADEESTUDIOS')

    celdas_en_negrita = Hoja1["A1":"D1"]
    fuente = Font(bold=True)

    for row in celdas_en_negrita:
        for cell in row:
            cell.font = fuente

    for row in celdas_en_negrita:
        for cell in row:
            cell.font = fuente
    book.save("filtro01.xlsx")

# ----------FIN PRIMER FILTRO 01-------------------------------------------------------------------------
# -----------INICIO DEL SEGUNDO FILTRO---------------------------------------------------------------


def eliminar_asteriscos():
    import pandas as pd
    # Abre el archivo para eliminar los asteriscos de la columna ENTIDAD FEDERATIVA

    df = pd.read_excel(io="filtro01.xlsx", sheet_name="Hoja1")

    fd1 = df.drop(df[df['ENTIDAD FEDERATIVA'] == '***'].index)
    print(fd1)
    fd1.to_excel('filtro02.xlsx')

    df = pd.read_excel(io="filtro02.xlsx")

    fd2 = df.drop(df[df['MUNICIPIO'] == '***'].index)
    print(fd2)
    fd2.to_excel('filtro02.xlsx')

    df = pd.read_excel(io="filtro02.xlsx")

    fd3 = df.drop(df[df['INSTITUCIÓN DE EDUCACIÓN SUPERIOR'] == '***'].index)
    print(fd3)
    fd3.to_excel('filtro02.xlsx')

    df = pd.read_excel(io="filtro02.xlsx")

    fd4 = df.drop(df.columns[[0, 1, 2]], axis='columns')
    print(fd4)
    fd4.to_excel('filtro02.xlsx')

# -----------FIN DEL SEGUNDO FILTRO 02-----------------------------------------------------------

# ---------INICIO DEL FILTRO TRES 03-----------------------------------------------------------


def filtro_areas():

    df = pd.read_excel('filtro02.xlsx')

    # Filtrar por Areas
    filtro = df[df.PROGRAMADEESTUDIOS.str.contains('|'.join(searchfor))]
    # Imprimir el primer resultado
    print(filtro)
    print('Aqui ya termino el filtro')
    filtro.to_excel("Filtro03.xlsx", index=False)


def filtro_clasificacion():

    # Abrir el archivo Filtro con la variable df2
    df2 = pd.read_excel('Filtro03.xlsx')
    conditionlist = [
        (df2['INSTITUCIÓN DE EDUCACIÓN SUPERIOR'].str.contains(
            '|'.join(FiltroOTROS))),
        (df2['INSTITUCIÓN DE EDUCACIÓN SUPERIOR'] .str.contains(
            '|'.join(FiltroPOLITECNICA))),
        (df2['INSTITUCIÓN DE EDUCACIÓN SUPERIOR'] .str.contains(
            '|'.join(FiltroTECNM))),
        (df2['INSTITUCIÓN DE EDUCACIÓN SUPERIOR'] .str.contains(
            '|'.join(FiltroTECNOLOGIA))),
        (df2['INSTITUCIÓN DE EDUCACIÓN SUPERIOR'] .str.contains('|'.join(FiltroUNIVERSIDAD)))]
    choicelist = ['OTRO', 'POLITECNICA', 'TECNM', 'TECNOLOGICA', 'UNIVERSIDAD']
    # Aqui se agrega una nueva columna llamada Clasificación
    clasificacion = df2['Clasificación'] = np.select(
        conditionlist, choicelist, default='No especificado')

    # Se guarda El archivo con el nombre de Data incluyendo el primer filtro de Areas y Tipo de institución
    df2.to_excel("Data.xlsx", index=False)
    # Imprimir el segundo resultado y Final
    print(df2)


def filtro_AreaCarrera():

    df = pd.read_excel('Data.xlsx')

    conditionlist = [
        (df['PROGRAMADEESTUDIOS'].str.contains(
            '|'.join(FiltroAreasComputacional))),
        (df['PROGRAMADEESTUDIOS'].str.contains('|'.join(FiltroAreasSoftware))),
        (df['PROGRAMADEESTUDIOS'].str.contains(
            '|'.join(FiltroAreasInformatica))),
        (df['PROGRAMADEESTUDIOS'].str.contains(
            '|'.join(FiltroAreasInteligencia))),
        (df['PROGRAMADEESTUDIOS'].str.contains('|'.join(FiltroAreasOtros))),
        (df['PROGRAMADEESTUDIOS'].str.contains(
            '|'.join(FiltroAreasSistemasCom))),
        (df['PROGRAMADEESTUDIOS'].str.contains(
            '|'.join(FiltroAreasSistemasInf))),
        (df['PROGRAMADEESTUDIOS'].str.contains(
            '|'.join(FiltroAreasTecnologias))),
        (df['PROGRAMADEESTUDIOS'].str.contains(
            '|'.join(FiltroAreasTecDeLaInformacion))),
        (df['PROGRAMADEESTUDIOS'].str.contains(
            '|'.join(FiltroAreasTelematica))),
    ]
    choicelist = ['CIENCIAS COMPUTACIONALES', 'DESARROLLO DE SOFTWARE', 'INFORMÁTICA', 'INTELIGENCIA ARTIFICIAL', 'OTROS',
                  'SISTEMAS COMPUTACIONALES', 'SISTEMAS DE INFORMACIÓN', 'TECNOLOGÍAS DE INFORMACIÓN', 'TECNOLOGIAS DE LA INFORMACIÓN', 'TELEMÁTICA']
    clasificacion = df['Carrera Areas'] = np.select(
        conditionlist, choicelist, default='No especificado')

    df.to_excel("Data.xlsx", index=False)
    print(df)

    print("------------------Fin del filtrado----------------")


# Lees la imagen:

# He colocado ruta relativa, es decir, la imagen a la misma

# altura de la aplicación. Si prefieres, puedes colocar una

# ruta absoluta.


# --------------------------------------------------------------------------------------------------------
imagen = PhotoImage(file="Imagenes/logo.png")

fondo = Label(ventana, image=imagen).place(x=660, y=0)

# .............................. .......................................

"""label2 = Label(canvas1, text="INSTITUTO TECNOLÓGICO DE CHETUMAL",

               font=("Times New Roman", 20), bg="#d4d9e3")

label2.pack()
label2.place(x=240, y=75)"""


# --------1 CANVAS CON LOS BOTONES Y SUS RESPECTIVOS FUNCIONES---------------------------------------------

c = Canvas(ventana, width=230, height=290, bg="#92B4EC")
c.place(x=240, y=290)
c.create_text(115, 50, text="01",

              fill="black", font=('Fine  30 bold'))

c.create_text(110, 130, text="Cargar el archivo de excel a cargar",

              fill="black", font=('Fine  8 bold'))

button1 = Button(ventana, text="Cargar", width=25,
                 height=2, command=abrir_archivo)

button1.pack(pady=25)
button1.place(x=265, y=520)

# --------2 CANVAS CON LOS BOTONES Y SUS RESPECTIVOS FUNCIONES---------------------------------------------

indica = Label(ventana, fg='white', bg='#004b60',

               text='Ubicación del archivo', font=('Arial', 7, 'bold'))

indica.place(x=155, y=600)
# -----------CANVAS 2----------------------------------------------------------------------------------------
c2 = Canvas(ventana, width=230, height=290, bg="#FD5F00")
c2.place(x=520, y=290)

c2.create_text(120, 50, text="02",

               fill="black", font=('Fine  30 bold'))


button2 = Button(ventana, text="Filtrar Datos",

                 width=25, height=2, command=lambda: [datos_excel(), municipio_excel(), institucion_excel(), programa_excel(), eliminar_asteriscos(), filtro_areas(), filtro_clasificacion(), filtro_AreaCarrera()])
button2.pack(pady=20)
button2.place(x=545, y=520)

# -------------------PARA QUE FUNCIONE LA VENTANA ES EL MAINLOOP()---------------------------------------------------


ventana.mainloop()
